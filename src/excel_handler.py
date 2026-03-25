"""Excel lesen und schreiben mit openpyxl."""

import os
from pathlib import Path
from typing import Iterator, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from utils import logger

# Spalten-Mapping (1-basiert)
COL_ISIN = 1            # A
COL_GROUP_INVESTMENT = 2  # B
COL_MS_SEGMENTIERUNG = 3  # C (Morningstar)
COL_CLAUDE_SEG = 7      # G
COL_FONDSTYP = 8        # H
COL_ANLEGERTYP = 9      # I
COL_KUNDENTYP = 10      # J
COL_KONFIDENZ = 11      # K
COL_PDF_DATEINAME = 12  # L
COL_STATUS = 13         # M

HEADER_ROW = 1
DATA_START_ROW = 2

# Farben für Status-Anzeige
COLOR_OK = "C6EFCE"       # Grün
COLOR_FEHLER = "FFC7CE"   # Rot
COLOR_UNKLAR = "FFEB9C"   # Gelb


def load_workbook_safe(excel_path: str):
    """Lädt eine Excel-Datei (mit Fehlerbehandlung)."""
    path = Path(excel_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel-Datei nicht gefunden: {excel_path}")

    try:
        wb = openpyxl.load_workbook(excel_path)
        logger.info(f"Excel geladen: {path.name}")
        return wb
    except Exception as e:
        raise RuntimeError(f"Fehler beim Laden von {path.name}: {e}")


def ensure_output_columns(ws) -> None:
    """Fügt Spaltenüberschriften für die neuen Spalten hinzu, falls fehlend."""
    headers = {
        COL_CLAUDE_SEG: "Claude_Segmentierung",
        COL_FONDSTYP: "Fondstyp",
        COL_ANLEGERTYP: "Anlegertyp",
        COL_KUNDENTYP: "Kundentyp",
        COL_KONFIDENZ: "Konfidenz",
        COL_PDF_DATEINAME: "PDF_Dateiname",
        COL_STATUS: "Verarbeitungsstatus",
    }

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="BDD7EE")  # Hellblau

    for col, name in headers.items():
        cell = ws.cell(row=HEADER_ROW, column=col)
        if not cell.value:
            cell.value = name
            cell.font = header_font
            cell.fill = header_fill


def iter_unprocessed_isins(excel_path: str, skip_done: bool = True) -> Iterator[tuple]:
    """
    Iterator über noch nicht verarbeitete ISINs.

    Yields:
        (row_number, isin, fund_name, ms_segmentierung)
    """
    wb = load_workbook_safe(excel_path)
    ws = wb.active

    ensure_output_columns(ws)

    count_total = 0
    count_skipped = 0

    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=False):
        row_num = row[0].row

        isin_cell = row[COL_ISIN - 1]
        isin = str(isin_cell.value or "").strip()

        if not isin or isin == "None":
            continue

        count_total += 1

        # Bereits verarbeitete überspringen
        if skip_done:
            status_cell = row[COL_STATUS - 1]
            status = str(status_cell.value or "").strip().lower()
            if status in ("ok", "fehler"):
                count_skipped += 1
                continue

        fund_name_cell = row[COL_GROUP_INVESTMENT - 1]
        fund_name = str(fund_name_cell.value or "").strip()

        ms_seg_cell = row[COL_MS_SEGMENTIERUNG - 1]
        ms_seg = str(ms_seg_cell.value or "").strip()

        yield row_num, isin, fund_name, ms_seg

    wb.close()
    logger.info(f"ISINs gesamt: {count_total}, übersprungen: {count_skipped}")


def count_total_rows(excel_path: str) -> int:
    """Zählt die Anzahl der ISIN-Zeilen."""
    wb = load_workbook_safe(excel_path)
    ws = wb.active
    count = 0
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        if row[COL_ISIN - 1]:
            count += 1
    wb.close()
    return count


def write_result(
    excel_path: str,
    row_num: int,
    result: dict,
    pdf_filename: str = "",
    status: str = "ok",
) -> None:
    """
    Schreibt das Klassifizierungsergebnis in die Excel-Datei.
    Speichert sofort (crash-sicher).
    """
    wb = load_workbook_safe(excel_path)
    ws = wb.active

    ensure_output_columns(ws)

    # Werte schreiben
    ws.cell(row=row_num, column=COL_CLAUDE_SEG).value = result.get("segmentierung", "")
    ws.cell(row=row_num, column=COL_FONDSTYP).value = result.get("fondstyp", "")
    ws.cell(row=row_num, column=COL_ANLEGERTYP).value = result.get("anlegertyp", "")
    ws.cell(row=row_num, column=COL_KUNDENTYP).value = result.get("kundentyp", "")
    ws.cell(row=row_num, column=COL_KONFIDENZ).value = result.get("konfidenz", "")
    ws.cell(row=row_num, column=COL_PDF_DATEINAME).value = pdf_filename
    ws.cell(row=row_num, column=COL_STATUS).value = status

    # Farbliche Hervorhebung
    status_cell = ws.cell(row=row_num, column=COL_STATUS)
    if status == "ok":
        seg = result.get("segmentierung", "unklar").lower()
        if seg == "unklar":
            status_cell.fill = PatternFill("solid", fgColor=COLOR_UNKLAR)
        else:
            status_cell.fill = PatternFill("solid", fgColor=COLOR_OK)
    elif status == "fehler":
        status_cell.fill = PatternFill("solid", fgColor=COLOR_FEHLER)

    try:
        wb.save(excel_path)
    except PermissionError:
        # Datei ist offen in Excel → Backup speichern
        backup_path = excel_path.replace(".xlsx", "_backup.xlsx")
        wb.save(backup_path)
        logger.warning(f"Excel gesperrt, gespeichert als: {backup_path}")
    finally:
        wb.close()


def write_error(excel_path: str, row_num: int, error_msg: str) -> None:
    """Schreibt einen Fehler in die Excel-Zeile."""
    result = {"segmentierung": "", "fondstyp": "", "anlegertyp": "", "kundentyp": "",
              "konfidenz": "", "begruendung": error_msg}
    write_result(excel_path, row_num, result, status="fehler")


def get_isin_row(excel_path: str, isin: str) -> Optional[int]:
    """Sucht die Zeilennummer einer ISIN."""
    wb = load_workbook_safe(excel_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=DATA_START_ROW):
        if str(row[COL_ISIN - 1].value or "").strip() == isin:
            wb.close()
            return row[0].row
    wb.close()
    return None


def adjust_column_widths(excel_path: str) -> None:
    """Passt die Spaltenbreiten an den Inhalt an."""
    wb = load_workbook_safe(excel_path)
    ws = wb.active

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    wb.save(excel_path)
    wb.close()
