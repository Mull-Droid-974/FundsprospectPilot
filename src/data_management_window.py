"""Datenverwaltungs-Fenster: ISIN-Import, Metadaten-Fetch, Ergebnis-Übersicht und Excel-Export."""

import queue
import threading
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

import results_store

BG_MAIN = "#1e1e2e"
BG_PANEL = "#2a2a3e"
BG_INPUT = "#313145"
FG_TEXT = "#cdd6f4"
FG_MUTED = "#7f849c"
ACCENT_BLUE = "#89b4fa"
ACCENT_GREEN = "#a6e3a1"
ACCENT_RED = "#f38ba8"
ACCENT_YELLOW = "#f9e2af"
ACCENT_LAVENDER = "#b4befe"
BTN_BG = "#45475a"
BTN_ACTIVE = "#585b70"


class DataManagementWindow(tk.Toplevel):
    """
    Datenverwaltungs-Fenster mit drei Tabs:
    - Import: ISIN-Grundmenge aus Excel laden
    - Metadaten: fundinfo-Metadaten für alle ISINs ohne subfonds_id laden
    - Ergebnisse: DB-Übersicht + Export
    """

    def __init__(self, parent: tk.Widget):
        super().__init__(parent)
        self.title("Datenverwaltung")
        self.configure(bg=BG_MAIN)
        self.geometry("720x560")
        self.minsize(600, 420)

        self._meta_worker = None
        self._meta_queue: queue.Queue = queue.Queue()

        self._build_ui()
        self.refresh_results()
        self._poll_meta_queue()

    def _build_ui(self):
        notebook = ttk.Notebook(self)
        notebook.pack(fill="both", expand=True, padx=12, pady=12)

        # Tab 1: Import
        self._import_tab = tk.Frame(notebook, bg=BG_MAIN)
        notebook.add(self._import_tab, text="  📥  ISIN-Import  ")
        self._build_import_tab(self._import_tab)

        # Tab 2: Metadaten
        self._meta_tab = tk.Frame(notebook, bg=BG_MAIN)
        notebook.add(self._meta_tab, text="  🔍  Metadaten  ")
        self._build_meta_tab(self._meta_tab)

        # Tab 3: Ergebnisse
        self._results_tab = tk.Frame(notebook, bg=BG_MAIN)
        notebook.add(self._results_tab, text="  📊  Ergebnisse  ")
        self._build_results_tab(self._results_tab)

    # ── Import-Tab ────────────────────────────────────────────────────────────

    def _build_import_tab(self, parent):
        self._current_mapping: dict[str, str] = {}   # excel_header → db_field
        self._header_to_idx:   dict[str, int]  = {}   # excel_header → col index

        tk.Label(
            parent,
            text="Importiert ISINs aus einer Excel-Datei. Spalten-Mapping wird nach Dateiauswahl konfiguriert.",
            bg=BG_MAIN, fg=FG_MUTED, font=("Segoe UI", 9),
            justify="left", anchor="w", wraplength=660,
        ).pack(anchor="w", padx=16, pady=(12, 6))

        # ── Dateiauswahl ──────────────────────────────────────────────────────
        file_row = tk.Frame(parent, bg=BG_MAIN)
        file_row.pack(fill="x", padx=16, pady=(0, 4))

        tk.Label(file_row, text="Excel-Datei:", bg=BG_MAIN, fg=FG_MUTED,
                 font=("Segoe UI", 9)).pack(side="left")
        self.var_import_path = tk.StringVar()
        tk.Entry(
            file_row, textvariable=self.var_import_path,
            bg=BG_INPUT, fg=FG_TEXT, insertbackground=FG_TEXT,
            font=("Segoe UI", 9), relief="flat", bd=4, width=46,
        ).pack(side="left", padx=(8, 4), fill="x", expand=True)
        tk.Button(
            file_row, text="...", command=self._browse_import,
            bg=BTN_BG, fg=FG_TEXT, relief="flat",
            font=("Segoe UI", 8), padx=6, cursor="hand2",
        ).pack(side="left")

        # ── Mapping-Status / Vorschau ─────────────────────────────────────────
        self._import_info = tk.Label(
            parent, text="", bg=BG_MAIN, fg=FG_MUTED,
            font=("Segoe UI", 9), anchor="w", wraplength=660,
        )
        self._import_info.pack(anchor="w", padx=16, pady=(2, 0))

        self._mapping_summary = tk.Label(
            parent, text="", bg=BG_MAIN, fg=ACCENT_BLUE,
            font=("Segoe UI", 8), anchor="w", wraplength=660,
        )
        self._mapping_summary.pack(anchor="w", padx=16, pady=(0, 4))

        # ── Aktions-Buttons ───────────────────────────────────────────────────
        action_row = tk.Frame(parent, bg=BG_MAIN)
        action_row.pack(anchor="w", padx=16, pady=(0, 4))

        self._btn_import = tk.Button(
            action_row, text="  ▶  Jetzt importieren  ",
            command=self._run_import,
            bg="#1e3a1e", fg=ACCENT_GREEN, relief="flat",
            font=("Segoe UI", 10, "bold"), padx=14, pady=6,
            cursor="hand2", state="disabled",
        )
        self._btn_import.pack(side="left")

        self._btn_remap = tk.Button(
            action_row, text="  ⚙  Spalten neu konfigurieren  ",
            command=self._open_mapping_dialog,
            bg=BTN_BG, fg=ACCENT_BLUE, relief="flat",
            font=("Segoe UI", 9), padx=10, pady=6,
            cursor="hand2", state="disabled",
        )
        self._btn_remap.pack(side="left", padx=(8, 0))

        # ── Log ───────────────────────────────────────────────────────────────
        self._import_log = tk.Text(
            parent, bg=BG_PANEL, fg=FG_TEXT,
            font=("Consolas", 8), relief="flat",
            state="disabled", wrap="word", height=10,
        )
        self._import_log.pack(fill="both", expand=True, padx=16, pady=12)

    # ── Import-Logik ──────────────────────────────────────────────────────────

    def _browse_import(self):
        path = filedialog.askopenfilename(
            title="Excel-Datei wählen",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Alle Dateien", "*.*")],
        )
        if not path:
            return
        self.var_import_path.set(path)

        # Spaltenköpfe + Zeilenzahl lesen
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            headers: list[str] = []
            if ws:
                first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
                headers = [str(c).strip() if c is not None else "" for c in first]
            row_count = max(0, ws.max_row - 1) if ws and ws.max_row else 0
            wb.close()
        except Exception as exc:
            self._import_info.config(text=f"Fehler beim Lesen: {exc}", fg=ACCENT_RED)
            return

        self._import_info.config(
            text=f"Datei: {Path(path).name}  |  ~{row_count} Zeilen  |  {len(headers)} Spalten erkannt",
            fg=ACCENT_BLUE,
        )
        self._header_to_idx = {h: i for i, h in enumerate(headers)}

        # Mapping-Dialog öffnen
        from excel_mapping_dialog import ExcelMappingDialog, load_saved_mapping
        saved = load_saved_mapping()
        dlg = ExcelMappingDialog(self, headers, saved)
        self.wait_window(dlg)

        if dlg.result:
            self._current_mapping = dlg.result
            self._update_mapping_summary()
            self._btn_import.config(state="normal")
            self._btn_remap.config(state="normal")

    def _open_mapping_dialog(self):
        """Mapping für die bereits gewählte Datei neu konfigurieren."""
        path = self.var_import_path.get().strip()
        if not path:
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            headers: list[str] = []
            if ws:
                first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
                headers = [str(c).strip() if c is not None else "" for c in first]
            wb.close()
        except Exception as exc:
            messagebox.showerror("Fehler", str(exc), parent=self)
            return

        from excel_mapping_dialog import ExcelMappingDialog, load_saved_mapping
        saved = load_saved_mapping()
        dlg = ExcelMappingDialog(self, headers, saved)
        self.wait_window(dlg)
        if dlg.result:
            self._current_mapping = dlg.result
            self._update_mapping_summary()

    def _update_mapping_summary(self):
        if not self._current_mapping:
            self._mapping_summary.config(text="")
            return
        parts = [f"{h} → {f}" for h, f in self._current_mapping.items()]
        self._mapping_summary.config(
            text="Mapping:  " + "  |  ".join(parts)
        )

    def _run_import(self):
        path = self.var_import_path.get().strip()
        if not path:
            messagebox.showerror("Fehler", "Bitte Excel-Datei wählen.", parent=self)
            return
        if not self._current_mapping:
            messagebox.showerror("Fehler", "Bitte zuerst Spalten konfigurieren.", parent=self)
            return

        self._log_import("Starte Import...", clear=True)
        mapping = dict(self._current_mapping)
        header_to_idx = dict(self._header_to_idx)

        def worker():
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                rows_data: list[dict] = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict: dict[str, str] = {}
                    for header, db_field in mapping.items():
                        idx = header_to_idx.get(header)
                        if idx is not None and idx < len(row) and row[idx] is not None:
                            val = str(row[idx]).strip()
                            if val and val != "None":
                                row_dict[db_field] = val
                    if row_dict.get("isin"):
                        rows_data.append(row_dict)
                wb.close()

                self.after(0, lambda: self._log_import(
                    f"{len(rows_data)} ISINs gelesen, importiere..."
                ))
                imported, skipped = results_store.import_base_set(rows_data)
                self.after(0, lambda: self._log_import(
                    f"✅ Fertig: {imported} neu importiert, {skipped} aktualisiert."
                ))
                self.after(0, self.refresh_results)

            except Exception as exc:
                self.after(0, lambda: self._log_import(f"❌ Fehler: {exc}"))

        threading.Thread(target=worker, daemon=True).start()

    def _log_import(self, msg: str, clear: bool = False):
        self._import_log.config(state="normal")
        if clear:
            self._import_log.delete("1.0", "end")
        self._import_log.insert("end", msg + "\n")
        self._import_log.see("end")
        self._import_log.config(state="disabled")

    # ── Metadaten-Tab ─────────────────────────────────────────────────────────

    def _build_meta_tab(self, parent):
        tk.Label(
            parent,
            text=(
                "Lädt Umbrella-ID, Subfonds-Name, Anteilsklasse und Prospekt-URL\n"
                "von fundinfo.com für alle ISINs, die noch keine Metadaten haben."
            ),
            bg=BG_MAIN, fg=FG_MUTED, font=("Segoe UI", 9),
            justify="left", anchor="w",
        ).pack(anchor="w", padx=16, pady=(12, 6))

        self._meta_stats = tk.Label(
            parent, text="", bg=BG_MAIN, fg=ACCENT_BLUE, font=("Segoe UI", 9), anchor="w"
        )
        self._meta_stats.pack(anchor="w", padx=16, pady=(0, 6))

        # Aktions-Buttons
        action_row = tk.Frame(parent, bg=BG_MAIN)
        action_row.pack(anchor="w", padx=16, pady=(0, 4))

        self._btn_meta = tk.Button(
            action_row, text="  🔍  Metadaten laden  ",
            command=self._run_meta_fetch,
            bg="#1e2e3a", fg=ACCENT_BLUE, relief="flat",
            font=("Segoe UI", 10, "bold"), padx=14, pady=6,
            cursor="hand2",
        )
        self._btn_meta.pack(side="left")

        self._btn_meta_stop = tk.Button(
            action_row, text="  ⏹  Stopp  ",
            command=self._stop_meta_worker,
            bg="#3a2000", fg=ACCENT_YELLOW, relief="flat",
            font=("Segoe UI", 9), padx=10, pady=6,
            cursor="hand2", state="disabled",
        )
        self._btn_meta_stop.pack(side="left", padx=(8, 0))

        tk.Button(
            action_row, text="  ↺  Nicht-gef. zurücksetzen  ",
            command=self._reset_not_found,
            bg=BTN_BG, fg=ACCENT_RED, relief="flat",
            font=("Segoe UI", 9), padx=10, pady=6,
            cursor="hand2",
        ).pack(side="left", padx=(8, 0))

        self._meta_force_var = tk.BooleanVar(value=False)
        tk.Checkbutton(
            action_row, text="Alle neu laden",
            variable=self._meta_force_var,
            bg=BG_MAIN, fg=FG_MUTED, selectcolor=BG_INPUT,
            activebackground=BG_MAIN, activeforeground=FG_TEXT,
            font=("Segoe UI", 9), cursor="hand2",
        ).pack(side="left", padx=(16, 4))

        tk.Label(
            action_row, text="Parallel:", bg=BG_MAIN, fg=FG_MUTED, font=("Segoe UI", 9)
        ).pack(side="left", padx=(8, 4))
        self._meta_parallel_var = tk.IntVar(value=4)
        tk.Spinbox(
            action_row, from_=1, to=16, textvariable=self._meta_parallel_var,
            width=3, bg=BG_INPUT, fg=FG_TEXT, insertbackground=FG_TEXT,
            buttonbackground=BTN_BG, relief="flat", font=("Segoe UI", 9),
        ).pack(side="left")

        # Fortschrittsbalken
        prog_row = tk.Frame(parent, bg=BG_MAIN)
        prog_row.pack(fill="x", padx=16, pady=(4, 0))

        self._meta_prog_var = tk.DoubleVar(value=0)
        self._meta_prog = ttk.Progressbar(
            prog_row, variable=self._meta_prog_var, maximum=100, length=320
        )
        self._meta_prog.pack(side="left", padx=(0, 10))

        self._meta_status = tk.Label(
            prog_row, text="", bg=BG_MAIN, fg=FG_MUTED, font=("Segoe UI", 8)
        )
        self._meta_status.pack(side="left")

        # Log
        self._meta_log = tk.Text(
            parent, bg=BG_PANEL, fg=FG_TEXT,
            font=("Consolas", 8), relief="flat",
            state="disabled", wrap="word", height=10,
        )
        self._meta_log.pack(fill="both", expand=True, padx=16, pady=12)

        self._refresh_meta_stats()

    def _refresh_meta_stats(self):
        try:
            queue_rows = results_store.get_meta_queue()
            total = len(results_store.get_all_results())
            n = len(queue_rows)
            if n == 0:
                self._meta_stats.config(
                    text=f"✅ Alle {total} ISINs haben Metadaten.", fg=ACCENT_GREEN
                )
            else:
                self._meta_stats.config(
                    text=f"{n} von {total} ISINs ohne Metadaten (subfonds_id leer).", fg=ACCENT_BLUE
                )
        except Exception:
            pass

    def _reset_not_found(self):
        try:
            n = results_store.reset_not_found()
        except Exception as exc:
            messagebox.showerror("Fehler", str(exc), parent=self)
            return
        if n == 0:
            messagebox.showinfo("Kein Reset nötig", "Keine als 'nicht gefunden' markierten ISINs.", parent=self)
        else:
            messagebox.showinfo("Zurückgesetzt", f"{n} ISIN(s) zurückgesetzt — werden beim nächsten Lauf erneut abgefragt.", parent=self)
            self._refresh_meta_stats()

    def _run_meta_fetch(self):
        from prospekt_worker import ProspektWorker
        from pathlib import Path as _Path

        force = self._meta_force_var.get()
        try:
            if force:
                results_store.reset_not_found()
                rows = results_store.get_all_results()
            else:
                rows = results_store.get_meta_queue()
        except Exception as exc:
            messagebox.showerror("Fehler", str(exc), parent=self)
            return

        if not rows:
            messagebox.showinfo("Fertig", "Alle ISINs haben bereits Metadaten.", parent=self)
            return

        self._meta_queue = queue.Queue()
        pdf_folder = _Path(__file__).parent.parent / "data" / "prospekte"
        parallel = max(1, min(16, self._meta_parallel_var.get()))
        self._meta_worker = ProspektWorker(
            rows, pdf_folder, self._meta_queue,
            meta_parallel=parallel, phase1_only=True, force_reload=force,
        )
        self._meta_worker.start()
        self._btn_meta.config(state="disabled")
        self._btn_meta_stop.config(state="normal")
        self._meta_prog_var.set(0)
        self._meta_status.config(text=f"0 / {len(rows)}")
        self._log_meta(f"Starte Metadaten-Abruf für {len(rows)} ISIN(s) …", clear=True)

    def _stop_meta_worker(self):
        if self._meta_worker:
            self._meta_worker.stop()

    def _poll_meta_queue(self):
        try:
            while True:
                from prospekt_worker import ProspektEvent
                evt: ProspektEvent = self._meta_queue.get_nowait()
                self._handle_meta_event(evt)
        except Exception:
            pass
        self.after(200, self._poll_meta_queue)

    def _handle_meta_event(self, evt):
        if evt.type == "log":
            self._log_meta(f"[{evt.isin}] {evt.message}" if evt.isin else evt.message)
        elif evt.type in ("progress", "error"):
            prefix = "✓" if evt.type == "progress" else "✗"
            self._log_meta(f"{prefix} [{evt.isin}] {evt.message}")
            if evt.total > 0:
                pct = (evt.done + evt.failed) / evt.total * 100
                self._meta_prog_var.set(pct)
                self._meta_status.config(
                    text=f"{evt.done + evt.failed}/{evt.total}  ✓{evt.done}  ✗{evt.failed}"
                )
        elif evt.type == "done":
            self._log_meta(f"✅ {evt.message}")
            self._meta_status.config(text=evt.message)
            self._meta_prog_var.set(100)
            self._btn_meta.config(state="normal")
            self._btn_meta_stop.config(state="disabled")
            self._refresh_meta_stats()
            self.refresh_results()

    def _log_meta(self, msg: str, clear: bool = False):
        self._meta_log.config(state="normal")
        if clear:
            self._meta_log.delete("1.0", "end")
        self._meta_log.insert("end", msg + "\n")
        self._meta_log.see("end")
        self._meta_log.config(state="disabled")

    # ── Ergebnisse-Tab ────────────────────────────────────────────────────────

    def _build_results_tab(self, parent):
        # Statistik-Zeile
        self._stats_label = tk.Label(
            parent, text="", bg=BG_MAIN, fg=FG_MUTED, font=("Segoe UI", 9), anchor="w"
        )
        self._stats_label.pack(anchor="w", padx=16, pady=(10, 4))

        # Treeview
        cols = ("isin", "fondsname", "segmentierung", "konfidenz", "analysiert_am")
        headers = ("ISIN", "Fondsname", "Segmentierung", "Konfidenz", "Analysiert am")
        widths = (130, 260, 110, 80, 130)

        tree_frame = tk.Frame(parent, bg=BG_MAIN)
        tree_frame.pack(fill="both", expand=True, padx=16, pady=(0, 4))

        self._tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=14)
        for col, hdr, w in zip(cols, headers, widths):
            self._tree.heading(col, text=hdr)
            self._tree.column(col, width=w, minwidth=60)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self._tree.yview)
        self._tree.configure(yscrollcommand=vsb.set)
        self._tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="left", fill="y")

        # Button-Leiste
        btn_row = tk.Frame(parent, bg=BG_MAIN)
        btn_row.pack(fill="x", padx=16, pady=(0, 12))

        tk.Button(
            btn_row, text="↺  Aktualisieren", command=self.refresh_results,
            bg=BTN_BG, fg=FG_TEXT, relief="flat",
            font=("Segoe UI", 9), padx=10, pady=4, cursor="hand2"
        ).pack(side="left")

        tk.Button(
            btn_row, text="📤  Excel-Export", command=self._export_excel,
            bg=BTN_BG, fg=ACCENT_BLUE, relief="flat",
            font=("Segoe UI", 9), padx=10, pady=4, cursor="hand2"
        ).pack(side="left", padx=(6, 0))

        tk.Button(
            btn_row, text="🗑  Eintrag löschen", command=self._delete_selected,
            bg=BTN_BG, fg=ACCENT_RED, relief="flat",
            font=("Segoe UI", 9), padx=10, pady=4, cursor="hand2"
        ).pack(side="right")

    def refresh_results(self):
        """Lädt DB-Daten neu und aktualisiert die Treeview und Metadaten-Stats."""
        self._refresh_meta_stats()
        try:
            rows = results_store.get_all_results()
            stats = results_store.get_stats()

            self._stats_label.config(
                text=(
                    f"Gesamt: {stats['total']}  |  "
                    f"Retail: {stats['retail']}  |  "
                    f"Institutional: {stats['institutional']}  |  "
                    f"Unklar: {stats['unklar']}"
                )
            )

            self._tree.delete(*self._tree.get_children())
            for r in rows:
                self._tree.insert("", "end", values=(
                    r.get("isin", ""),
                    r.get("fondsname", ""),
                    r.get("segmentierung", ""),
                    r.get("konfidenz", ""),
                    r.get("analysiert_am", ""),
                ))
        except Exception:
            pass

    def _export_excel(self):
        path = filedialog.asksaveasfilename(
            title="Excel speichern",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not path:
            return
        try:
            results_store.export_to_excel(path)
            messagebox.showinfo("Export", f"Exportiert nach:\n{path}", parent=self)
        except Exception as e:
            messagebox.showerror("Fehler", str(e), parent=self)

    def _delete_selected(self):
        selected = self._tree.selection()
        if not selected:
            return
        isin = self._tree.item(selected[0])["values"][0]
        if not messagebox.askyesno("Löschen", f"ISIN {isin} wirklich löschen?", parent=self):
            return
        results_store.delete_result(str(isin))
        self.refresh_results()
