"""
Permanenter Ergebnis-Speicher für klassifizierte Fonds.

Speichert pro ISIN ein Klassifizierungsergebnis in einer SQLite-Datenbank
(data/output/results.db). Bei erneuter Analyse wird der Eintrag überschrieben.

Kein GUI-Code — reine Datenbankschicht.
"""

import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Optional

# ─── Pfad ────────────────────────────────────────────────────────────────────
_DB_PATH = Path(__file__).parent.parent / "data" / "output" / "results.db"

_COLUMNS = [
    "isin", "fund_id", "fondsname", "fondstyp", "anlegertyp", "kundentyp",
    "segmentierung", "pruef_segmentierung", "konfidenz", "begruendung",
    "quelle", "modell", "pdf_datei", "ter",
    "analysiert_am", "erstellt_am", "ueberschrieben_am",
    "prospekt_pfad", "prospekt_url",
]


# ─── Initialisierung ─────────────────────────────────────────────────────────

def _connect() -> sqlite3.Connection:
    _DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(str(_DB_PATH))
    con.row_factory = sqlite3.Row
    return con


def init_db():
    """Erstellt die Tabelle falls sie noch nicht existiert und migriert neue Spalten."""
    with _connect() as con:
        con.execute("""
            CREATE TABLE IF NOT EXISTS fund_results (
                isin                TEXT PRIMARY KEY,
                fund_id             TEXT DEFAULT '',
                fondsname           TEXT DEFAULT '',
                fondstyp            TEXT DEFAULT '',
                anlegertyp          TEXT DEFAULT '',
                kundentyp           TEXT DEFAULT '',
                segmentierung       TEXT DEFAULT '',
                pruef_segmentierung TEXT DEFAULT '',
                konfidenz           TEXT DEFAULT '',
                begruendung         TEXT DEFAULT '',
                quelle              TEXT DEFAULT '',
                modell              TEXT DEFAULT '',
                pdf_datei           TEXT DEFAULT '',
                ter                 TEXT DEFAULT '',
                analysiert_am       TEXT DEFAULT '',
                erstellt_am         TEXT DEFAULT '',
                ueberschrieben_am   TEXT DEFAULT ''
            )
        """)
        # Migration: neue Spalten zu bestehenden DBs hinzufügen
        for col_def in [
            "begruendung TEXT DEFAULT ''",
            "erstellt_am TEXT DEFAULT ''",
            "ueberschrieben_am TEXT DEFAULT ''",
            "fund_id TEXT DEFAULT ''",
            "pruef_segmentierung TEXT DEFAULT ''",
            "ter TEXT DEFAULT ''",
            "prospekt_pfad TEXT DEFAULT ''",
            "prospekt_url TEXT DEFAULT ''",
        ]:
            try:
                con.execute(f"ALTER TABLE fund_results ADD COLUMN {col_def}")
            except Exception:
                pass  # Spalte existiert bereits


# ─── Schreiben ────────────────────────────────────────────────────────────────

def upsert_result(
    isin: str,
    fondsname: str,
    result: dict,
    pdf_datei: str = "",
):
    """
    Speichert oder aktualisiert ein Klassifizierungsergebnis.

    result: Rückgabe-Dict von process_single_pdf() / classify_with_llm()
    """
    if not isin:
        return

    tokens = result.get("_tokens", {})
    modell = tokens.get("model", "") if tokens else ""
    quelle = result.get("_source", "")
    now = datetime.now().strftime("%Y-%m-%d %H:%M")

    with _connect() as con:
        existing = con.execute(
            "SELECT isin FROM fund_results WHERE isin = ?", (isin,)
        ).fetchone()

        if existing:
            con.execute("""
                UPDATE fund_results SET
                    fondsname         = ?,
                    fondstyp          = ?,
                    anlegertyp        = ?,
                    kundentyp         = ?,
                    segmentierung     = ?,
                    konfidenz         = ?,
                    begruendung       = ?,
                    quelle            = ?,
                    modell            = ?,
                    pdf_datei         = ?,
                    analysiert_am     = ?,
                    ueberschrieben_am = ?
                WHERE isin = ?
            """, (
                fondsname or "",
                result.get("fondstyp", "") or "",
                result.get("anlegertyp", "") or "",
                result.get("kundentyp", "") or "",
                result.get("segmentierung", "") or "",
                result.get("konfidenz", "") or "",
                result.get("begruendung", "") or "",
                quelle, modell,
                pdf_datei or "",
                now, now,
                isin,
            ))
        else:
            con.execute("""
                INSERT INTO fund_results
                    (isin, fondsname, fondstyp, anlegertyp, kundentyp,
                     segmentierung, konfidenz, begruendung, quelle, modell,
                     pdf_datei, analysiert_am, erstellt_am, ueberschrieben_am)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '')
            """, (
                isin,
                fondsname or "",
                result.get("fondstyp", "") or "",
                result.get("anlegertyp", "") or "",
                result.get("kundentyp", "") or "",
                result.get("segmentierung", "") or "",
                result.get("konfidenz", "") or "",
                result.get("begruendung", "") or "",
                quelle, modell,
                pdf_datei or "",
                now, now,
            ))


# ─── Lesen ────────────────────────────────────────────────────────────────────

def get_all_results() -> list[dict]:
    """Gibt alle Einträge sortiert nach Analysedatum (neueste zuerst) zurück."""
    init_db()
    with _connect() as con:
        rows = con.execute(
            "SELECT * FROM fund_results ORDER BY analysiert_am DESC"
        ).fetchall()
    return [dict(r) for r in rows]


def count() -> int:
    """Anzahl gespeicherter ISINs."""
    init_db()
    with _connect() as con:
        return con.execute("SELECT COUNT(*) FROM fund_results").fetchone()[0]


def get_stats() -> dict:
    """Gibt Zählungen nach Segmentierung zurück."""
    init_db()
    with _connect() as con:
        rows = con.execute(
            "SELECT segmentierung, COUNT(*) AS n FROM fund_results GROUP BY segmentierung"
        ).fetchall()
    counts = {r["segmentierung"]: r["n"] for r in rows}
    return {
        "retail":        counts.get("retail", 0),
        "institutional": counts.get("institutional", 0),
        "unklar":        counts.get("unklar", 0),
        "total":         sum(counts.values()),
    }


# ─── Löschen ─────────────────────────────────────────────────────────────────

def delete_result(isin: str):
    """Löscht einen Eintrag anhand der ISIN."""
    with _connect() as con:
        con.execute("DELETE FROM fund_results WHERE isin = ?", (isin,))


# ─── Excel-Export ─────────────────────────────────────────────────────────────

def export_to_excel(path: str):
    """Exportiert alle Einträge als Excel-Datei."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = get_all_results()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fonds-Ergebnisse"

    # Header
    headers = ["ISIN", "Fondsname", "Fondstyp", "Anlegertyp", "Kundentyp",
               "Segmentierung", "Konfidenz", "Begründung", "Quelle", "Modell",
               "PDF-Datei", "Analysiert am", "Erstellt am", "Überschrieben am"]
    keys    = _COLUMNS

    header_fill = PatternFill("solid", fgColor="1e1e2e")
    header_font = Font(bold=True, color="cdd6f4")

    for col_idx, (h, _) in enumerate(zip(headers, keys), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Daten
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(keys, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))

    # Spaltenbreiten
    widths = [16, 35, 12, 22, 22, 16, 10, 45, 8, 22, 40, 18, 18, 18]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = w

    wb.save(path)


# ─── Anreicherung ────────────────────────────────────────────────────────────

def update_enrichment(
    isin: str,
    fondstyp: str = "",
    anlegertyp: str = "",
    kundentyp: str = "",
    ter: str = "",
    overwrite_empty_only: bool = True,
):
    """
    Befüllt Anreicherungsfelder für eine ISIN.
    Wenn overwrite_empty_only=True, werden nur leere Felder überschrieben.
    Analyse-Felder (segmentierung, konfidenz, begruendung) bleiben immer unberührt.
    """
    if not isin:
        return

    with _connect() as con:
        if overwrite_empty_only:
            con.execute("""
                UPDATE fund_results SET
                    fondstyp   = CASE WHEN fondstyp   = '' OR fondstyp   IS NULL THEN ? ELSE fondstyp   END,
                    anlegertyp = CASE WHEN anlegertyp = '' OR anlegertyp IS NULL THEN ? ELSE anlegertyp END,
                    kundentyp  = CASE WHEN kundentyp  = '' OR kundentyp  IS NULL THEN ? ELSE kundentyp  END,
                    ter        = CASE WHEN ter        = '' OR ter        IS NULL THEN ? ELSE ter        END
                WHERE isin = ?
            """, (fondstyp, anlegertyp, kundentyp, ter, isin))
        else:
            con.execute("""
                UPDATE fund_results SET
                    fondstyp   = ?,
                    anlegertyp = ?,
                    kundentyp  = ?,
                    ter        = ?
                WHERE isin = ?
            """, (fondstyp, anlegertyp, kundentyp, ter, isin))


def get_enrichment_queue(limit: int = 0) -> list[dict]:
    """
    Gibt ISINs zurück bei denen mindestens ein Anreicherungsfeld leer ist.
    Sortiert nach fund_id (ISINs mit fund_id zuerst — bessere Fundinfo-Abdeckung).
    """
    init_db()
    query = """
        SELECT isin, fund_id, fondsname, pruef_segmentierung
        FROM fund_results
        WHERE fondstyp = '' OR fondstyp IS NULL
           OR anlegertyp = '' OR anlegertyp IS NULL
           OR kundentyp = '' OR kundentyp IS NULL
           OR ter = '' OR ter IS NULL
        ORDER BY CASE WHEN fund_id != '' THEN 0 ELSE 1 END, isin
    """
    if limit > 0:
        query += f" LIMIT {limit}"
    with _connect() as con:
        rows = con.execute(query).fetchall()
    return [dict(r) for r in rows]


# ─── Grundmenge importieren ───────────────────────────────────────────────────

_SEG_NORMALIZE = {
    "institutionell": "institutional",
    "institutional":  "institutional",
    "retail":         "retail",
    "privat":         "retail",
}


def import_base_set(rows: list[dict]) -> tuple[int, int]:
    """
    Importiert ISINs als Grundmenge — bestehende ISINs werden NICHT überschrieben.

    rows: Liste von Dicts mit den Schlüsseln:
          isin, fund_id, fondsname, pruef_segmentierung

    Returns:
        (n_imported, n_skipped)
    """
    init_db()
    imported = 0
    skipped  = 0
    now      = datetime.now().strftime("%Y-%m-%d %H:%M")

    with _connect() as con:
        for row in rows:
            isin = (row.get("isin") or "").strip()
            if not isin:
                continue

            raw_seg = (row.get("pruef_segmentierung") or "").strip().lower()
            pruef_seg = _SEG_NORMALIZE.get(raw_seg, raw_seg)

            existing = con.execute(
                "SELECT isin FROM fund_results WHERE isin = ?", (isin,)
            ).fetchone()

            if existing:
                # FundID, Fondsname und Prüfsegment immer überschreiben
                con.execute("""
                    UPDATE fund_results
                    SET fund_id             = ?,
                        fondsname           = ?,
                        pruef_segmentierung = ?
                    WHERE isin = ?
                """, (
                    (row.get("fund_id") or "").strip(),
                    (row.get("fondsname") or "").strip(),
                    pruef_seg,
                    isin,
                ))
                skipped += 1
            else:
                con.execute("""
                    INSERT INTO fund_results
                        (isin, fund_id, fondsname, pruef_segmentierung, erstellt_am)
                    VALUES (?, ?, ?, ?, ?)
                """, (
                    isin,
                    (row.get("fund_id") or "").strip(),
                    (row.get("fondsname") or "").strip(),
                    pruef_seg,
                    now,
                ))
                imported += 1

    return imported, skipped


# ─── Prospekt-Download ───────────────────────────────────────────────────────

def get_by_prospekt_url(url: str) -> Optional[dict]:
    """Gibt den ersten DB-Eintrag zurück, der diese Prospekt-URL bereits hat."""
    if not url:
        return None
    with _connect() as con:
        row = con.execute(
            "SELECT * FROM fund_results WHERE prospekt_url = ? AND prospekt_pfad != '' LIMIT 1",
            (url,)
        ).fetchone()
    return dict(row) if row else None


def update_prospekt(isin: str, prospekt_pfad: str, prospekt_url: str):
    """Speichert lokalen Dateipfad und Original-URL eines heruntergeladenen Prospekts."""
    if not isin:
        return
    with _connect() as con:
        con.execute(
            "UPDATE fund_results SET prospekt_pfad=?, prospekt_url=? WHERE isin=?",
            (prospekt_pfad or "", prospekt_url or "", isin),
        )


def get_result(isin: str) -> Optional[dict]:
    """Gibt einen einzelnen DB-Eintrag anhand der ISIN zurück."""
    init_db()
    with _connect() as con:
        row = con.execute(
            "SELECT * FROM fund_results WHERE isin = ?", (isin,)
        ).fetchone()
    return dict(row) if row else None


def get_prospekt_queue() -> list[dict]:
    """
    Gibt alle ISINs zurück, für die noch kein Prospekt vorliegt
    (prospekt_pfad leer ODER Datei nicht auf Disk vorhanden).
    """
    init_db()
    rows = get_all_results()
    return [
        r for r in rows
        if not r.get("prospekt_pfad") or not Path(r["prospekt_pfad"]).exists()
    ]


# DB beim Import initialisieren
init_db()
