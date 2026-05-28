"""
Ergebnis-Datenbank-Fenster.

Zeigt alle gespeicherten ISIN-Klassifizierungen in einer sortierbaren,
filterbaren Tabellenansicht. Wird von app.py als Toplevel geöffnet.
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import results_store

# ─── Farben ──────────────────────────────────────────────────────────────────
BG_MAIN     = "#1e1e2e"
BG_PANEL    = "#2a2a3e"
BG_INPUT    = "#313145"
FG_TEXT     = "#cdd6f4"
FG_MUTED    = "#7f849c"
FG_DIM      = "#6c7086"
ACCENT_BLUE     = "#89b4fa"
ACCENT_GREEN    = "#a6e3a1"
ACCENT_RED      = "#f38ba8"
ACCENT_YELLOW   = "#f9e2af"
ACCENT_LAVENDER = "#b4befe"
BTN_BG      = "#45475a"
BTN_ACTIVE  = "#585b70"

# Spalten-Konfiguration: (db-key, Anzeige-Header, Breite)
_COLS = [
    ("isin",          "ISIN",          130),
    ("fondsname",     "Fondsname",     220),
    ("fondstyp",      "Fondstyp",      100),
    ("anlegertyp",    "Anlegertyp",    160),
    ("kundentyp",      "Kundentyp",     150),
    ("mindestanlage",  "Mindestanlage", 130),
    ("segmentierung",  "Segmentierung", 110),
    ("konfidenz",     "Konfidenz",      80),
    ("analysiert_am", "Analysiert am", 130),
    ("prospekt_pfad",    "Prospekt-Datei",  200),
    ("prospekt_url",     "Prospekt-URL",    180),
    ("subfonds_name",    "Unterfonds",      200),
    ("anteilsklasse",    "Anteilsklasse",   150),
    ("ausschuettungsart","Ausschüttung",     90),
    ("fondswaehrung",    "Währung",          70),
    ("fundinfo_ter",           "TER (API)",        80),
    ("fundinfo_investor_type", "Inv.Type (API)",    130),
    ("umbrella_id",            "Umbrella-ID",       150),
    ("ongoing_charges_datum",  "OC-Datum",           90),
    ("qualif_anleger_ch",           "Qual.Anleger CH",    90),
    ("institutional_ch",           "Institutional CH",   90),
    ("llm_segmentierung",             "LLM-Segment.",      110),
    ("llm_segmentierung_begruendung", "LLM-Begründung",    280),
    ("fondstyp_roh",                  "Fondstyp (Roh)",    180),
    ("anlegertyp_roh",                "Anlegertyp (Roh)",  180),
    ("kundentyp_roh",                 "Kundentyp (Roh)",   180),
]
_COL_KEYS = [c[0] for c in _COLS]

# Segmentierungs-Farben
_SEG_COLORS = {
    "retail":        ("#a6e3a1", "#1e2e1e"),
    "institutional": ("#89b4fa", "#1e2030"),
    "unklar":        ("#f9e2af", "#2e2a1e"),
}


class ResultsWindow(tk.Toplevel):
    """
    Permanentes Ergebnis-Fenster.

    Liest Daten aus results_store (SQLite) und zeigt sie als Treeview.
    Wird über app.py geöffnet; kann mehrfach geöffnet sein.
    """

    def __init__(self, parent: tk.Widget):
        super().__init__(parent)
        self.title("Ergebnis-Datenbank")
        self.configure(bg=BG_MAIN)
        self.geometry("1100x600")
        self.minsize(800, 400)

        self._all_rows: list[dict] = []      # Alle DB-Einträge (ungefiltert)
        self._visible_rows: list[dict] = []  # Aktuell angezeigte Zeilen (nach Filter)
        self._sort_col: str = "analysiert_am"
        self._sort_rev: bool = True

        self._build_ui()
        self.refresh()

    # ── UI-Aufbau ──────────────────────────────────────────────────────────

    def _build_ui(self):
        # ── Toolbar ──────────────────────────────────────────────
        toolbar = tk.Frame(self, bg=BG_PANEL)
        toolbar.pack(fill="x")

        inner = tk.Frame(toolbar, bg=BG_PANEL)
        inner.pack(fill="x", padx=12, pady=8)

        tk.Label(
            inner, text="Ergebnis-Datenbank",
            bg=BG_PANEL, fg=ACCENT_LAVENDER,
            font=("Segoe UI", 12, "bold"),
        ).pack(side="left")

        # Rechte Buttons
        tk.Button(
            inner, text="Excel exportieren",
            command=self._export_excel,
            bg=BTN_BG, fg=FG_TEXT, relief="flat",
            font=("Segoe UI", 9), padx=10, pady=3, cursor="hand2",
            activebackground=BTN_ACTIVE,
        ).pack(side="right", padx=(4, 0))

        tk.Button(
            inner, text="Ohne Metadaten exportieren",
            command=self._export_no_metadata,
            bg=BTN_BG, fg=ACCENT_YELLOW, relief="flat",
            font=("Segoe UI", 9), padx=10, pady=3, cursor="hand2",
            activebackground=BTN_ACTIVE,
        ).pack(side="right", padx=(4, 0))

        tk.Button(
            inner, text="Aktualisieren",
            command=self.refresh,
            bg=BTN_BG, fg=FG_TEXT, relief="flat",
            font=("Segoe UI", 9), padx=10, pady=3, cursor="hand2",
            activebackground=BTN_ACTIVE,
        ).pack(side="right", padx=(4, 0))

        # ── Suchleiste ────────────────────────────────────────────
        search_bar = tk.Frame(self, bg=BG_MAIN)
        search_bar.pack(fill="x", padx=12, pady=(8, 4))

        tk.Label(
            search_bar, text="🔍", bg=BG_MAIN, fg=FG_MUTED,
            font=("Segoe UI", 10),
        ).pack(side="left")

        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", lambda *_: self._apply_filter())
        tk.Entry(
            search_bar, textvariable=self._search_var,
            bg=BG_INPUT, fg=FG_TEXT, insertbackground=FG_TEXT,
            relief="flat", font=("Segoe UI", 10),
        ).pack(side="left", fill="x", expand=True, padx=(6, 0), ipady=4)

        self._count_var = tk.StringVar(value="")
        tk.Label(
            search_bar, textvariable=self._count_var,
            bg=BG_MAIN, fg=FG_MUTED, font=("Segoe UI", 9),
        ).pack(side="right", padx=(8, 0))

        # ── Treeview ─────────────────────────────────────────────
        tree_frame = tk.Frame(self, bg=BG_MAIN)
        tree_frame.pack(fill="both", expand=True, padx=12, pady=(0, 4))

        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "Results.Treeview",
            background=BG_PANEL,
            foreground=FG_TEXT,
            fieldbackground=BG_PANEL,
            rowheight=26,
            borderwidth=0,
            font=("Segoe UI", 9),
        )
        style.configure(
            "Results.Treeview.Heading",
            background=BG_INPUT,
            foreground=ACCENT_LAVENDER,
            relief="flat",
            font=("Segoe UI", 9, "bold"),
        )
        style.map("Results.Treeview",
                  background=[("selected", "#3d3d5c")],
                  foreground=[("selected", FG_TEXT)])
        style.map("Results.Treeview.Heading",
                  background=[("active", BTN_ACTIVE)])

        col_ids = [c[0] for c in _COLS]
        self._tree = ttk.Treeview(
            tree_frame,
            columns=col_ids,
            show="headings",
            style="Results.Treeview",
            selectmode="browse",
        )

        # Spalten konfigurieren
        for key, header, width in _COLS:
            self._tree.heading(
                key, text=header,
                command=lambda k=key: self._sort_by(k),
            )
            self._tree.column(key, width=width, minwidth=60, anchor="w")

        # Scrollbar
        vsb = ttk.Scrollbar(tree_frame, orient="vertical",
                             command=self._tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal",
                             command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self._tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        self._tree.bind("<Double-1>", self._on_double_click)

        # Tags für Segmentierungsfarben
        for seg, (fg, bg) in _SEG_COLORS.items():
            self._tree.tag_configure(seg, foreground=fg, background=bg)
        self._tree.tag_configure("alt", background="#252538")

        # ── Untere Aktionsleiste ──────────────────────────────────
        action_bar = tk.Frame(self, bg=BG_MAIN)
        action_bar.pack(fill="x", padx=12, pady=(0, 8))

        tk.Button(
            action_bar, text="Ausgewählten Eintrag löschen",
            command=self._delete_selected,
            bg=BTN_BG, fg=ACCENT_RED, relief="flat",
            font=("Segoe UI", 9), padx=10, pady=3, cursor="hand2",
            activebackground=BTN_ACTIVE,
        ).pack(side="left")

    # ── Daten laden / anzeigen ─────────────────────────────────────────────

    def refresh(self):
        """Liest alle Einträge aus der DB und zeigt sie an."""
        self._all_rows = results_store.get_all_results()
        self._apply_filter()

    def _apply_filter(self):
        """Filtert die angezeigten Zeilen nach dem Suchbegriff."""
        query = self._search_var.get().strip().lower()

        if query:
            rows = [
                r for r in self._all_rows
                if query in (r.get("isin") or "").lower()
                or query in (r.get("fondsname") or "").lower()
                or query in (r.get("fondstyp") or "").lower()
                or query in (r.get("segmentierung") or "").lower()
            ]
        else:
            rows = self._all_rows

        # Sortieren
        rows = sorted(
            rows,
            key=lambda r: (r.get(self._sort_col) or "").lower(),
            reverse=self._sort_rev,
        )

        self._visible_rows = rows
        self._fill_tree(rows)

        total = len(self._all_rows)
        shown = len(rows)
        if query:
            self._count_var.set(f"{shown} von {total} ISINs")
        else:
            self._count_var.set(f"{total} ISINs gespeichert")

    def _fill_tree(self, rows: list[dict]):
        self._tree.delete(*self._tree.get_children())
        for i, row in enumerate(rows):
            values = [row.get(k, "") or "" for k in _COL_KEYS]
            seg = (row.get("segmentierung") or "").lower()
            tag = seg if seg in _SEG_COLORS else ("alt" if i % 2 else "")
            self._tree.insert("", "end", iid=row["isin"], values=values, tags=(tag,))

    # ── Sortieren ──────────────────────────────────────────────────────────

    def _sort_by(self, col_key: str):
        if self._sort_col == col_key:
            self._sort_rev = not self._sort_rev
        else:
            self._sort_col = col_key
            self._sort_rev = False
        self._apply_filter()

        # Pfeil im Header anzeigen
        for key, header, _ in _COLS:
            arrow = (" ↑" if not self._sort_rev else " ↓") if key == col_key else ""
            self._tree.heading(key, text=header + arrow)

    # ── Aktionen ───────────────────────────────────────────────────────────

    def _on_double_click(self, event):
        sel = self._tree.selection()
        if not sel:
            return
        isin = sel[0]
        row = next((r for r in self._all_rows if r.get("isin") == isin), None)
        if row:
            self._show_detail(row)

    def _show_detail(self, row: dict):
        isin = row.get("isin", "")
        try:
            import morningstar_store
            ms = morningstar_store.get_morningstar(isin) or {}
            row = {**row, **ms}
        except Exception:
            pass

        win = tk.Toplevel(self)
        win.transient(self)
        win.grab_set()
        win.title(f"Detail — {row.get('isin', '')}  {row.get('fondsname', '')}")
        win.configure(bg=BG_MAIN)
        win.geometry("680x640")
        win.minsize(500, 400)

        # Titel
        tk.Label(
            win,
            text=f"{row.get('isin', '')}  —  {row.get('fondsname', '') or '—'}",
            bg=BG_PANEL, fg=ACCENT_LAVENDER,
            font=("Segoe UI", 11, "bold"),
            anchor="w", padx=14, pady=8,
        ).pack(fill="x")

        # Scrollbarer Inhaltsbereich
        canvas = tk.Canvas(win, bg=BG_MAIN, highlightthickness=0)
        vsb = ttk.Scrollbar(win, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        inner = tk.Frame(canvas, bg=BG_MAIN)
        canvas_win = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_resize(e):
            canvas.itemconfig(canvas_win, width=e.width)
        canvas.bind("<Configure>", _on_resize)

        def _on_frame_configure(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        inner.bind("<Configure>", _on_frame_configure)

        def _on_mousewheel(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        win.bind("<Destroy>", lambda e: canvas.unbind_all("<MouseWheel>"))

        # Feldgruppen
        _GROUPS = [
            ("Identifikation", ["isin", "fondsname", "subfonds_name", "anteilsklasse",
                                "ausschuettungsart", "fondswaehrung"]),
            ("Analyse", ["segmentierung", "fondstyp", "anlegertyp", "kundentyp",
                         "mindestanlage", "konfidenz", "analysiert_am"]),
            ("LLM-Segmentierung", ["llm_segmentierung", "llm_segmentierung_begruendung"]),
            ("Fundinfo API", ["fundinfo_ter", "fundinfo_investor_type", "umbrella_id",
                              "ongoing_charges_datum", "qualif_anleger_ch", "institutional_ch"]),
            ("Morningstar", ["ms_category", "ms_legal_type", "ms_investor_type",
                             "ms_mifid_category", "ms_rating", "ms_share_class_status",
                             "ms_inception_date", "ms_termination_date"]),
            ("Prospekt", ["prospekt_url", "prospekt_pfad"]),
        ]

        _LABELS = {k: h for k, h, _ in _COLS}
        try:
            from morningstar_client import AVAILABLE_DATAPOINTS as _MS_DP
            _LABELS.update({k: v[1] for k, v in _MS_DP.items()})
        except Exception:
            pass

        _ROH_MAP = {
            "fondstyp":      "fondstyp_roh",
            "anlegertyp":    "anlegertyp_roh",
            "kundentyp":     "kundentyp_roh",
            "mindestanlage": "mindestanlage_roh",
        }

        for group_title, keys in _GROUPS:
            # Gruppenheader
            tk.Label(
                inner, text=group_title,
                bg=BG_MAIN, fg=ACCENT_BLUE,
                font=("Segoe UI", 9, "bold"),
                anchor="w", padx=14, pady=6,
            ).pack(fill="x", pady=(10, 0))

            sep = tk.Frame(inner, bg=FG_DIM, height=1)
            sep.pack(fill="x", padx=14, pady=(0, 4))

            for key in keys:
                val = str(row.get(key) or "—")
                label = _LABELS.get(key, key)
                roh_key = _ROH_MAP.get(key)
                roh_val = str(row.get(roh_key) or "") if roh_key else ""

                row_frame = tk.Frame(inner, bg=BG_PANEL)
                row_frame.pack(fill="x", padx=14, pady=1)

                tk.Label(
                    row_frame, text=label,
                    bg=BG_PANEL, fg=FG_MUTED,
                    font=("Segoe UI", 8),
                    width=22, anchor="nw",
                ).pack(side="left", padx=(8, 4), pady=4)

                val_frame = tk.Frame(row_frame, bg=BG_PANEL)
                val_frame.pack(side="left", fill="x", expand=True, padx=(0, 8), pady=2)

                is_long = len(val) > 80 or "\n" in val
                if is_long:
                    txt = tk.Text(
                        val_frame, bg=BG_PANEL, fg=FG_TEXT,
                        font=("Segoe UI", 9),
                        relief="flat", wrap="word",
                        height=min(6, val.count("\n") + len(val) // 60 + 1),
                        bd=0,
                    )
                    txt.insert("1.0", val)
                    txt.configure(state="disabled")
                    txt.pack(fill="x", pady=(2, 0))
                else:
                    tk.Label(
                        val_frame, text=val,
                        bg=BG_PANEL, fg=FG_TEXT,
                        font=("Segoe UI", 9),
                        anchor="nw", justify="left",
                    ).pack(fill="x", pady=(2, 0))

                if roh_val:
                    roh_is_long = len(roh_val) > 80 or "\n" in roh_val
                    if roh_is_long:
                        roh_txt = tk.Text(
                            val_frame, bg=BG_PANEL, fg=FG_DIM,
                            font=("Segoe UI", 7),
                            relief="flat", wrap="word",
                            height=min(3, roh_val.count("\n") + len(roh_val) // 70 + 1),
                            bd=0,
                        )
                        roh_txt.insert("1.0", roh_val)
                        roh_txt.configure(state="disabled")
                        roh_txt.pack(fill="x", pady=(0, 2))
                    else:
                        tk.Label(
                            val_frame, text=roh_val,
                            bg=BG_PANEL, fg=FG_DIM,
                            font=("Segoe UI", 7),
                            anchor="nw", justify="left",
                        ).pack(fill="x", pady=(0, 2))

        # Schliessen-Button
        tk.Button(
            win, text="Schliessen",
            command=win.destroy,
            bg=BTN_BG, fg=FG_TEXT, relief="flat",
            font=("Segoe UI", 9), padx=14, pady=5, cursor="hand2",
            activebackground=BTN_ACTIVE,
        ).pack(pady=(6, 10))

    def _delete_selected(self):
        sel = self._tree.selection()
        if not sel:
            return
        isin = sel[0]
        if messagebox.askyesno(
            "Eintrag löschen",
            f"ISIN {isin} wirklich aus der Datenbank entfernen?",
            parent=self,
        ):
            results_store.delete_result(isin)
            self.refresh()

    def _export_excel(self):
        rows = self._visible_rows
        is_filtered = bool(self._search_var.get().strip())
        label = f"gefiltert, {len(rows)} Zeilen" if is_filtered else f"{len(rows)} Zeilen"

        path = filedialog.asksaveasfilename(
            parent=self,
            defaultextension=".xlsx",
            filetypes=[("Excel-Datei", "*.xlsx")],
            initialfile="fonds_ergebnisse.xlsx",
            title=f"Ergebnisse exportieren ({label})",
        )
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Fonds-Ergebnisse"

            header_fill = PatternFill("solid", fgColor="1e1e2e")
            header_font = Font(bold=True, color="cdd6f4")

            for col_idx, (_, header, _) in enumerate(_COLS, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for row_idx, row in enumerate(rows, start=2):
                for col_idx, (key, _, _) in enumerate(_COLS, start=1):
                    ws.cell(row=row_idx, column=col_idx,
                            value=row.get(key, "") or "")

            for col_idx, (_, _, px_width) in enumerate(_COLS, start=1):
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(col_idx)
                ].width = max(8, min(60, px_width // 7))

            wb.save(path)
            messagebox.showinfo(
                "Export erfolgreich",
                f"{len(rows)} Einträge ({label}) gespeichert:\n{path}",
                parent=self,
            )
        except Exception as e:
            messagebox.showerror("Export fehlgeschlagen", str(e), parent=self)

    def _export_no_metadata(self):
        rows = results_store.get_no_metadata_results()
        if not rows:
            messagebox.showinfo(
                "Keine Einträge",
                "Alle ISINs in der Datenbank haben einen Prospekt-URL.",
                parent=self,
            )
            return

        path = filedialog.asksaveasfilename(
            parent=self,
            defaultextension=".xlsx",
            filetypes=[("Excel-Datei", "*.xlsx"), ("CSV-Datei", "*.csv")],
            initialfile="isins_ohne_metadaten.xlsx",
            title=f"ISINs ohne Metadaten exportieren ({len(rows)} Einträge)",
        )
        if not path:
            return

        _EXPORT_COLS = [
            ("isin",                   "ISIN"),
            ("fondsname",              "Fondsname"),
            ("subfonds_name",          "Subfonds"),
            ("anteilsklasse",          "Anteilsklasse"),
            ("ausschuettungsart",      "Ausschüttung"),
            ("fondswaehrung",          "Währung"),
            ("erstellt_am",            "Importiert am"),
            ("prospekt_nicht_gefunden","Prospekt-Suche (Datum)"),
        ]

        try:
            if path.lower().endswith(".csv"):
                import csv
                with open(path, "w", newline="", encoding="utf-8-sig") as f:
                    writer = csv.writer(f, delimiter=";")
                    writer.writerow([h for _, h in _EXPORT_COLS])
                    for row in rows:
                        writer.writerow([row.get(k, "") or "" for k, _ in _EXPORT_COLS])
            else:
                import openpyxl
                from openpyxl.styles import Alignment, Font, PatternFill
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Ohne Metadaten"

                header_fill = PatternFill("solid", fgColor="2e2800")
                header_font = Font(bold=True, color="f9e2af")
                col_widths   = [18, 40, 35, 22, 14, 10, 18, 22]

                for ci, (_, header) in enumerate(_EXPORT_COLS, start=1):
                    cell = ws.cell(row=1, column=ci, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                    ws.column_dimensions[
                        openpyxl.utils.get_column_letter(ci)
                    ].width = col_widths[ci - 1]

                for ri, row in enumerate(rows, start=2):
                    for ci, (key, _) in enumerate(_EXPORT_COLS, start=1):
                        ws.cell(row=ri, column=ci, value=row.get(key, "") or "")

                wb.save(path)

            messagebox.showinfo(
                "Export erfolgreich",
                f"{len(rows)} ISINs ohne Metadaten gespeichert:\n{path}",
                parent=self,
            )
        except Exception as e:
            messagebox.showerror("Export fehlgeschlagen", str(e), parent=self)

    # ── Öffentliche Methode für app.py ─────────────────────────────────────

    def add_result(self, isin: str, fondsname: str, result: dict, pdf_datei: str = ""):
        """
        Speichert ein neues Ergebnis und aktualisiert die Anzeige sofort.
        Wird von app.py nach jeder Analyse aufgerufen.
        """
        results_store.upsert_result(isin, fondsname, result, pdf_datei)
        self.refresh()
